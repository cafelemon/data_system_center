from datetime import UTC, datetime
from io import BytesIO
from typing import Annotated, Any
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session, selectinload
from starlette import status
from starlette.responses import StreamingResponse

from app.api.deps import get_current_user, get_db, require_admin_user
from app.core.responses import ApiResponse, ok
from app.models import (
    Archive,
    ArchiveStatus,
    ArchiveType,
    Department,
    RetentionPeriod,
    User,
)
from app.schemas.archive import (
    ArchiveBatchDeletePayload,
    ArchiveListResponse,
    ArchiveOptionsResponse,
    ArchiveMedium,
    ArchivePayload,
    ArchiveRead,
    RetentionPeriodOption,
    SelectOption,
)
from app.services.operation_logs import write_operation_log

router = APIRouter(prefix="/archives", tags=["archives"])


def archive_query():
    return select(Archive).options(
        selectinload(Archive.archive_type),
        selectinload(Archive.status),
        selectinload(Archive.retention_period),
        selectinload(Archive.department),
    )


def archive_filter_conditions(
    *,
    keyword: str | None = None,
    archive_medium: ArchiveMedium | None = None,
    internal_archive_type: str | None = None,
    archive_type_id: int | None = None,
    department_id: int | None = None,
    status_id: int | None = None,
) -> list[Any]:
    conditions = [Archive.deleted_at.is_(None)]
    if keyword and keyword.strip():
        keyword_like = f"%{keyword.strip()}%"
        conditions.append(
            or_(
                Archive.archive_no.ilike(keyword_like),
                Archive.title.ilike(keyword_like),
            )
        )
    if archive_medium is not None:
        conditions.append(Archive.archive_medium == archive_medium)
    if internal_archive_type and internal_archive_type.strip():
        conditions.append(
            Archive.internal_archive_type.ilike(f"%{internal_archive_type.strip()}%")
        )
    if archive_type_id is not None:
        conditions.append(Archive.archive_type_id == archive_type_id)
    if department_id is not None:
        conditions.append(Archive.department_id == department_id)
    if status_id is not None:
        conditions.append(Archive.status_id == status_id)
    return conditions


def archive_filter_summary(
    *,
    keyword: str | None = None,
    archive_medium: ArchiveMedium | None = None,
    internal_archive_type: str | None = None,
    archive_type_id: int | None = None,
    department_id: int | None = None,
    status_id: int | None = None,
) -> str:
    parts = []
    if keyword and keyword.strip():
        parts.append(f"关键词={keyword.strip()}")
    if archive_medium is not None:
        parts.append(f"台账={archive_medium}")
    if internal_archive_type and internal_archive_type.strip():
        parts.append(f"内部档案类型={internal_archive_type.strip()}")
    if archive_type_id is not None:
        parts.append(f"档案类型ID={archive_type_id}")
    if department_id is not None:
        parts.append(f"部门ID={department_id}")
    if status_id is not None:
        parts.append(f"状态ID={status_id}")
    return "，".join(parts) if parts else "全部档案"


def normalize_optional_text(value: str | None) -> str | None:
    if value is None:
        return None
    stripped = value.strip()
    return stripped or None


def archive_values(payload: ArchivePayload) -> dict[str, Any]:
    archive_year = payload.archive_year
    if archive_year is None and payload.archive_date is not None:
        archive_year = payload.archive_date.year

    return {
        "archive_no": payload.archive_no.strip(),
        "title": payload.title.strip(),
        "archive_medium": payload.archive_medium,
        "archive_type_id": payload.archive_type_id,
        "internal_archive_type": payload.internal_archive_type.strip(),
        "status_id": payload.status_id,
        "retention_period_id": payload.retention_period_id,
        "department_id": payload.department_id,
        "paper_copies": payload.paper_copies
        if payload.archive_medium == "paper"
        else 0,
        "archive_date": payload.archive_date,
        "paper_storage_location": normalize_optional_text(payload.paper_storage_location)
        if payload.archive_medium == "paper"
        else None,
        "electronic_storage_path": normalize_optional_text(
            payload.electronic_storage_path
        )
        if payload.archive_medium == "electronic"
        else None,
        "archiver_name": normalize_optional_text(payload.archiver_name),
        "owner_name": normalize_optional_text(payload.owner_name),
        "archive_year": archive_year,
        "security_level": normalize_optional_text(payload.security_level),
        "importance_level": normalize_optional_text(payload.importance_level),
        "project_name": normalize_optional_text(payload.project_name),
        "related_party": normalize_optional_text(payload.related_party),
        "contract_no": normalize_optional_text(payload.contract_no),
        "keywords": normalize_optional_text(payload.keywords),
        "remarks": normalize_optional_text(payload.remarks),
    }


def get_archive_or_404(db: Session, archive_id: int) -> Archive:
    archive = db.scalar(
        archive_query().where(
            Archive.id == archive_id,
            Archive.deleted_at.is_(None),
        )
    )
    if archive is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="档案不存在",
        )
    return archive


def ensure_archive_no_unique(
    db: Session,
    archive_no: str,
    exclude_archive_id: int | None = None,
) -> None:
    query = select(Archive).where(Archive.archive_no == archive_no.strip())
    if exclude_archive_id is not None:
        query = query.where(Archive.id != exclude_archive_id)

    duplicate = db.scalar(query)
    if duplicate is not None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="档案编号已存在",
        )


def ensure_lookup_exists(
    db: Session,
    model: type[Any],
    item_id: int,
    label: str,
    *,
    require_enabled: bool = False,
    allowed_disabled_id: int | None = None,
) -> None:
    item = db.get(model, item_id)
    if item is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{label}不存在",
        )
    if (
        require_enabled
        and not item.enabled
        and item_id != allowed_disabled_id
    ):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{label}已停用",
        )


def validate_archive_lookups(
    db: Session,
    payload: ArchivePayload,
    archive: Archive | None = None,
) -> None:
    ensure_lookup_exists(
        db,
        ArchiveType,
        payload.archive_type_id,
        "档案类型",
        require_enabled=True,
        allowed_disabled_id=archive.archive_type_id if archive else None,
    )
    ensure_lookup_exists(db, ArchiveStatus, payload.status_id, "档案状态")
    ensure_lookup_exists(
        db,
        RetentionPeriod,
        payload.retention_period_id,
        "保管期限",
        require_enabled=True,
        allowed_disabled_id=archive.retention_period_id if archive else None,
    )
    ensure_lookup_exists(db, Department, payload.department_id, "归档部门")


def archive_export_headers(archive_medium: ArchiveMedium) -> list[str]:
    common = [
        "档案编号",
        "档案名称",
        "档案类型",
        "内部档案类型",
        "状态",
        "保管期限",
        "归档部门",
        "归档人",
        "归档日期",
    ]
    if archive_medium == "paper":
        return [
            "档案编号",
            "档案名称",
            "纸质份数",
            "档案类型",
            "内部档案类型",
            "状态",
            "保管期限",
            "归档部门",
            "归档人",
            "归档日期",
            "存放位置",
            "责任人",
        ]
    return [*common, "存放路径", "责任人"]


def archive_export_row(archive: Archive, archive_medium: ArchiveMedium) -> list[Any]:
    if archive_medium == "paper":
        return [
            archive.archive_no,
            archive.title,
            archive.paper_copies,
            archive.archive_type.name,
            archive.internal_archive_type,
            archive.status.name,
            archive.retention_period.name,
            archive.department.name,
            archive.archiver_name or "",
            archive.archive_date,
            archive.paper_storage_location or "",
            archive.owner_name or "",
        ]
    return [
        archive.archive_no,
        archive.title,
        archive.archive_type.name,
        archive.internal_archive_type,
        archive.status.name,
        archive.retention_period.name,
        archive.department.name,
        archive.archiver_name or "",
        archive.archive_date,
        archive.electronic_storage_path or "",
        archive.owner_name or "",
    ]


def write_archive_log(
    db: Session,
    *,
    operator: User,
    operation_type: str,
    archive: Archive,
    detail: str,
    request: Request,
) -> None:
    write_operation_log(
        db,
        module="档案管理",
        operation_type=operation_type,
        operator=operator,
        target_id=str(archive.id),
        target_name=archive.title,
        detail=detail,
        request=request,
    )


@router.get("/options", response_model=ApiResponse[ArchiveOptionsResponse])
def get_archive_options(
    db: Session = Depends(get_db),
    _current_user: User = Depends(get_current_user),
) -> ApiResponse[ArchiveOptionsResponse]:
    archive_types = db.scalars(
        select(ArchiveType).order_by(ArchiveType.sort_order.asc(), ArchiveType.id.asc())
    ).all()
    statuses = db.scalars(
        select(ArchiveStatus).order_by(
            ArchiveStatus.sort_order.asc(),
            ArchiveStatus.id.asc(),
        )
    ).all()
    departments = db.scalars(
        select(Department).order_by(
            Department.enabled.desc(),
            Department.sort_order.asc(),
            Department.id.asc(),
        )
    ).all()
    retention_periods = db.scalars(
        select(RetentionPeriod).order_by(
            RetentionPeriod.sort_order.asc(),
            RetentionPeriod.id.asc(),
        )
    ).all()

    return ok(
        ArchiveOptionsResponse(
            archive_types=[
                SelectOption(
                    id=item.id,
                    name=item.name,
                    code=item.code,
                    enabled=item.enabled,
                )
                for item in archive_types
            ],
            statuses=[
                SelectOption(
                    id=item.id,
                    name=item.name,
                    code=item.code,
                    enabled=item.enabled,
                )
                for item in statuses
            ],
            departments=[
                SelectOption(
                    id=item.id,
                    name=item.name,
                    code=item.code,
                    enabled=item.enabled,
                )
                for item in departments
            ],
            retention_periods=[
                RetentionPeriodOption(
                    id=item.id,
                    name=item.name,
                    years=item.years,
                    enabled=item.enabled,
                )
                for item in retention_periods
            ],
        )
    )


@router.get("", response_model=ApiResponse[ArchiveListResponse])
def list_archives(
    keyword: str | None = Query(default=None, max_length=100),
    archive_medium: ArchiveMedium = Query(default="paper"),
    internal_archive_type: str | None = Query(default=None, max_length=100),
    archive_type_id: int | None = Query(default=None),
    department_id: int | None = Query(default=None),
    status_id: int | None = Query(default=None),
    page: Annotated[int, Query(ge=1)] = 1,
    page_size: Annotated[int, Query(ge=1, le=100)] = 10,
    db: Session = Depends(get_db),
    _current_user: User = Depends(get_current_user),
) -> ApiResponse[ArchiveListResponse]:
    conditions = archive_filter_conditions(
        keyword=keyword,
        archive_medium=archive_medium,
        internal_archive_type=internal_archive_type,
        archive_type_id=archive_type_id,
        department_id=department_id,
        status_id=status_id,
    )

    total = db.scalar(select(func.count()).select_from(Archive).where(*conditions))
    archives = db.scalars(
        archive_query()
        .where(*conditions)
        .order_by(Archive.archive_no.asc(), Archive.id.asc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).all()

    return ok(
        ArchiveListResponse(
            items=[ArchiveRead.model_validate(archive) for archive in archives],
            total=total or 0,
            page=page,
            page_size=page_size,
        )
    )


@router.get("/export")
def export_archives(
    request: Request,
    keyword: str | None = Query(default=None, max_length=100),
    archive_medium: ArchiveMedium = Query(default="paper"),
    internal_archive_type: str | None = Query(default=None, max_length=100),
    archive_type_id: int | None = Query(default=None),
    department_id: int | None = Query(default=None),
    status_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> StreamingResponse:
    conditions = archive_filter_conditions(
        keyword=keyword,
        archive_medium=archive_medium,
        internal_archive_type=internal_archive_type,
        archive_type_id=archive_type_id,
        department_id=department_id,
        status_id=status_id,
    )
    archives = db.scalars(
        archive_query()
        .where(*conditions)
        .order_by(Archive.archive_no.asc(), Archive.id.asc())
    ).all()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "档案目录"
    worksheet.freeze_panes = "A2"

    headers = archive_export_headers(archive_medium)
    worksheet.append(headers)

    header_fill = PatternFill("solid", fgColor="EAF0FF")
    header_font = Font(bold=True, color="111827")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for archive in archives:
        worksheet.append(archive_export_row(archive, archive_medium))
        if archive.archive_date:
            date_column = 10 if archive_medium == "paper" else 9
            worksheet.cell(
                row=worksheet.max_row,
                column=date_column,
            ).number_format = "yyyy-mm-dd"

    column_widths = (
        [18, 30, 12, 14, 18, 12, 12, 18, 14, 14, 18, 12]
        if archive_medium == "paper"
        else [18, 30, 14, 18, 12, 12, 18, 14, 14, 30, 12]
    )
    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    filter_summary = archive_filter_summary(
        keyword=keyword,
        archive_medium=archive_medium,
        internal_archive_type=internal_archive_type,
        archive_type_id=archive_type_id,
        department_id=department_id,
        status_id=status_id,
    )
    write_operation_log(
        db,
        module="档案管理",
        operation_type="导出Excel",
        operator=current_user,
        target_id="archives_export",
        target_name="档案目录导出",
        detail=f"导出档案目录 {len(archives)} 条，筛选条件：{filter_summary}",
        request=request,
    )
    db.commit()

    filename = f"档案目录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    quoted_filename = quote(filename)
    return StreamingResponse(
        stream,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quoted_filename}",
        },
    )


@router.delete("/batch", response_model=ApiResponse[ArchiveListResponse])
def batch_delete_archives(
    payload: ArchiveBatchDeletePayload,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_user),
) -> ApiResponse[ArchiveListResponse]:
    archive_ids = list(dict.fromkeys(payload.archive_ids))
    archives = db.scalars(
        archive_query().where(
            Archive.id.in_(archive_ids),
            Archive.deleted_at.is_(None),
        )
    ).all()

    now = datetime.now(UTC)
    for archive in archives:
        archive.deleted_at = now
        archive.updated_by = current_user.id

    write_operation_log(
        db,
        module="档案管理",
        operation_type="批量删除档案",
        operator=current_user,
        target_id=",".join(str(archive.id) for archive in archives),
        target_name="批量档案",
        detail=f"批量软删除档案 {len(archives)} 条",
        request=request,
    )
    deleted_reads = [ArchiveRead.model_validate(archive) for archive in archives]
    db.commit()
    return ok(
        ArchiveListResponse(
            items=deleted_reads,
            total=len(deleted_reads),
            page=1,
            page_size=len(deleted_reads),
        )
    )


@router.get("/{archive_id}", response_model=ApiResponse[ArchiveRead])
def get_archive(
    archive_id: int,
    db: Session = Depends(get_db),
    _current_user: User = Depends(get_current_user),
) -> ApiResponse[ArchiveRead]:
    return ok(ArchiveRead.model_validate(get_archive_or_404(db, archive_id)))


@router.post("", response_model=ApiResponse[ArchiveRead])
def create_archive(
    payload: ArchivePayload,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_user),
) -> ApiResponse[ArchiveRead]:
    values = archive_values(payload)
    ensure_archive_no_unique(db, values["archive_no"])
    validate_archive_lookups(db, payload)

    archive = Archive(
        **values,
        created_by=current_user.id,
        updated_by=current_user.id,
    )
    db.add(archive)
    db.flush()
    write_archive_log(
        db,
        operator=current_user,
        operation_type="新增档案",
        archive=archive,
        detail=f"新增档案 {archive.archive_no}",
        request=request,
    )
    db.commit()
    return ok(ArchiveRead.model_validate(get_archive_or_404(db, archive.id)))


@router.put("/{archive_id}", response_model=ApiResponse[ArchiveRead])
def update_archive(
    archive_id: int,
    payload: ArchivePayload,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_user),
) -> ApiResponse[ArchiveRead]:
    archive = get_archive_or_404(db, archive_id)
    values = archive_values(payload)
    ensure_archive_no_unique(db, values["archive_no"], exclude_archive_id=archive.id)
    validate_archive_lookups(db, payload, archive)

    for field, value in values.items():
        setattr(archive, field, value)
    archive.updated_by = current_user.id
    write_archive_log(
        db,
        operator=current_user,
        operation_type="编辑档案",
        archive=archive,
        detail=f"编辑档案 {archive.archive_no}",
        request=request,
    )
    db.commit()
    return ok(ArchiveRead.model_validate(get_archive_or_404(db, archive.id)))


@router.delete("/{archive_id}", response_model=ApiResponse[ArchiveRead])
def delete_archive(
    archive_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_user),
) -> ApiResponse[ArchiveRead]:
    archive = get_archive_or_404(db, archive_id)
    archive.deleted_at = datetime.now(UTC)
    archive.updated_by = current_user.id
    write_archive_log(
        db,
        operator=current_user,
        operation_type="软删除档案",
        archive=archive,
        detail=f"软删除档案 {archive.archive_no}",
        request=request,
    )
    archive_read = ArchiveRead.model_validate(archive)
    db.commit()
    return ok(archive_read)
